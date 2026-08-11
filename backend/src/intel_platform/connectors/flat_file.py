"""Flat file connector — CSV, Excel, JSON, JSONL ingestion through collection plans.

Supports:
- CSV/TSV with encoding detection, delimiter detection, header detection
- Excel (.xlsx) with sheet selection
- JSON / JSONL (nested → flattened)
- XML / KML (namespaces stripped, repeating record element auto-detected) —
  the shape most government feeds ship, including OFAC SDN and the EU
  consolidated sanctions list
- GeoJSON (one record per feature; geometry reduced to a representative point
  plus bounding box, because locations are stored as lat/lng scalars and there
  is no geometry column)

Security:
- CSV formula injection prevention (=, +, -, @, |, \\ prefixes sanitized)
- Row count limits to prevent memory exhaustion
- Column count limits
- JSON depth limits
- Encoding validation (whitelist of safe encodings)
- Excel workbook resource cleanup via context manager
- XML parsed with defusedxml, never the stdlib parser. Checked on Python 3.12:
  `xml.etree.ElementTree` refuses external entities but expands internal ones,
  so a billion-laughs document parses and inflates. defusedxml refuses both.
"""
from __future__ import annotations

import csv
import io
import json
import logging
from collections import Counter
from datetime import datetime
from typing import TYPE_CHECKING, Any

if TYPE_CHECKING:  # only for the XML element annotations below
    import xml.etree.ElementTree as ET

from intel_platform.connectors.base import (
    AcquireResult,
    ConnectorHealth,
    HealthStatus,
    SourceConnector,
    register_connector,
)

logger = logging.getLogger(__name__)

SUPPORTED_FORMATS = {"csv", "tsv", "xlsx", "xls", "json", "jsonl", "xml", "geojson"}
MAX_PREVIEW_ROWS = 50
MAX_PROFILE_UNIQUE = 100
MAX_ROWS = 500_000  # Hard limit: refuse files with more rows
MAX_COLUMNS = 1_000  # Hard limit: refuse files with more columns
MAX_JSON_SIZE = 100 * 1024 * 1024  # 100MB decoded text limit for JSON
# XML nesting deeper than this is a malformed or hostile document, not data.
MAX_XML_DEPTH = 40

# Characters that trigger formula execution in spreadsheet applications
_FORMULA_PREFIXES = ("=", "+", "-", "@", "|", "\\")

# Safe encodings whitelist — reject exotic encodings that could cause issues
_SAFE_ENCODINGS = {
    "utf-8", "utf-16", "utf-16-le", "utf-16-be", "utf-32",
    "ascii", "latin-1", "iso-8859-1", "iso-8859-2", "iso-8859-15",
    "windows-1250", "windows-1251", "windows-1252", "windows-1253", "windows-1254",
    "windows-1255", "windows-1256", "cp1252", "cp437", "cp850",
    "mac-roman", "euc-jp", "shift_jis", "iso-2022-jp",
    "euc-kr", "gb2312", "gbk", "gb18030", "big5",
    "koi8-r", "koi8-u",
}


def _sanitize_cell(value: str) -> str:
    """Sanitize a cell value to prevent CSV formula injection.

    Prefixes dangerous characters with a single quote to neutralize
    formula execution in spreadsheet applications.
    """
    if value and isinstance(value, str) and value.lstrip().startswith(_FORMULA_PREFIXES):
        return "'" + value
    return value


def _sanitize_header(name: str) -> str:
    """Sanitize a column header name."""
    sanitized = _sanitize_cell(name)
    # Also strip control characters from headers
    sanitized = "".join(c for c in sanitized if c.isprintable() or c in (" ", "\t"))
    return sanitized.strip() or "unnamed"


def _detect_encoding(raw: bytes) -> str:
    """Detect encoding using chardet, fallback to utf-8.

    Only accepts encodings from a safe whitelist to prevent
    encoding-based attacks.
    """
    try:
        import chardet
        result = chardet.detect(raw[:10000])
        detected = (result.get("encoding") or "utf-8").lower().replace("_", "-")
        confidence = result.get("confidence", 0)
        # Reject low-confidence detections and unsafe encodings
        if confidence < 0.5 or detected not in _SAFE_ENCODINGS:
            logger.warning("Rejected encoding %s (confidence=%.2f), using utf-8", detected, confidence)
            return "utf-8"
        return detected
    except ImportError:
        return "utf-8"


def _detect_delimiter(sample: str) -> str:
    """Detect CSV delimiter from sample text."""
    try:
        dialect = csv.Sniffer().sniff(sample[:8192], delimiters=",\t;|")
        return dialect.delimiter
    except csv.Error:
        return ","


def _detect_has_header(sample: str, delimiter: str) -> bool:
    """Detect if CSV has header row."""
    try:
        return csv.Sniffer().has_header(sample[:8192])
    except csv.Error:
        return True


def _infer_column_type(values: list[Any]) -> str:
    """Infer column type from sample values."""
    non_null = [v for v in values if v is not None and str(v).strip() != ""]
    if not non_null:
        return "null"

    # Try integer
    int_count = 0
    for v in non_null[:100]:
        try:
            int(str(v).replace(",", ""))
            int_count += 1
        except (ValueError, TypeError):
            break

    if int_count == len(non_null[:100]):
        return "integer"

    # Try float
    float_count = 0
    for v in non_null[:100]:
        try:
            float(str(v).replace(",", ""))
            float_count += 1
        except (ValueError, TypeError):
            break

    if float_count == len(non_null[:100]):
        return "float"

    # Try date
    from dateutil import parser as dateutil_parser
    date_count = 0
    for v in non_null[:20]:
        try:
            dateutil_parser.parse(str(v), fuzzy=False)
            date_count += 1
        except (ValueError, TypeError, OverflowError):
            break

    if date_count >= 15 or (date_count == len(non_null[:20]) and len(non_null) > 0):
        return "datetime"

    # Try boolean
    bool_vals = {"true", "false", "yes", "no", "1", "0", "t", "f", "y", "n"}
    if all(str(v).strip().lower() in bool_vals for v in non_null[:100]):
        return "boolean"

    return "string"


def _profile_column(name: str, values: list[Any]) -> dict:
    """Generate profiling stats for a single column."""
    total = len(values)
    nulls = sum(1 for v in values if v is None or str(v).strip() == "")
    non_null = [v for v in values if v is not None and str(v).strip() != ""]
    unique_vals = set(str(v) for v in non_null[:MAX_PROFILE_UNIQUE * 10])
    counter = Counter(str(v) for v in non_null[:1000])

    profile: dict = {
        "name": name,
        "total": total,
        "null_count": nulls,
        "null_rate": round(nulls / total, 4) if total else 0,
        "unique_count": min(len(unique_vals), MAX_PROFILE_UNIQUE),
        "top_values": [{"value": v, "count": c} for v, c in counter.most_common(10)],
    }

    # Numeric stats
    numeric_vals = []
    for v in non_null:
        try:
            numeric_vals.append(float(str(v).replace(",", "")))
        except (ValueError, TypeError):
            pass

    if numeric_vals and len(numeric_vals) > len(non_null) * 0.5:
        profile["min"] = min(numeric_vals)
        profile["max"] = max(numeric_vals)
        profile["mean"] = round(sum(numeric_vals) / len(numeric_vals), 4)

    return profile


def parse_csv(raw: bytes, config: dict) -> AcquireResult:
    """Parse CSV/TSV bytes into structured records with profiling."""
    encoding = config.get("encoding") or _detect_encoding(raw)
    text = raw.decode(encoding, errors="replace")

    delimiter = config.get("delimiter") or _detect_delimiter(text)
    has_header = config.get("has_header", _detect_has_header(text, delimiter))

    # Increase CSV field size limit to handle large cells (up to 10MB per field)
    csv.field_size_limit(10 * 1024 * 1024)

    try:
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)

        # Read rows with a hard limit to prevent memory exhaustion
        rows = []
        for i, row in enumerate(reader):
            if i > MAX_ROWS:
                return AcquireResult(
                    success=False,
                    error=f"File exceeds maximum row limit of {MAX_ROWS:,}. Truncate or split the file.")
            rows.append(row)
    except csv.Error as e:
        return AcquireResult(success=False, error=f"CSV parse error: {e}")

    if not rows:
        return AcquireResult(success=True, record_count=0)

    # Column count check
    if len(rows[0]) > MAX_COLUMNS:
        return AcquireResult(
            success=False,
            error=f"File has {len(rows[0])} columns, exceeding limit of {MAX_COLUMNS}.")

    if has_header:
        headers = [_sanitize_header(h.strip()) for h in rows[0]]
        data_rows = rows[1:]
    else:
        headers = [f"column_{i}" for i in range(len(rows[0]))]
        data_rows = rows

    # Build records with formula injection sanitization
    records = []
    for row_idx, row in enumerate(data_rows):
        record = {"_row_number": row_idx + 1}
        for col_idx, header in enumerate(headers):
            val = row[col_idx].strip() if col_idx < len(row) else None
            if val is not None:
                val = _sanitize_cell(val)
            record[header] = val
        records.append(record)

    # Schema inference
    column_values = {h: [r.get(h) for r in records] for h in headers}
    schema_info = {
        "columns": [
            {"name": h, "type": _infer_column_type(column_values[h]), "index": i}
            for i, h in enumerate(headers)
        ],
        "delimiter": delimiter,
        "encoding": encoding,
        "has_header": has_header,
    }

    # Profiling
    profiling = {
        "row_count": len(records),
        "column_count": len(headers),
        "columns": {h: _profile_column(h, column_values[h]) for h in headers},
    }

    # Preview rows
    preview_rows = records[:MAX_PREVIEW_ROWS]

    return AcquireResult(
        success=True,
        record_count=len(records),
        records=records,
        schema_info=schema_info,
        profiling=profiling,
        preview_rows=preview_rows,
        metadata={"format": "csv", "encoding": encoding, "delimiter": delimiter},
    )


def parse_excel(raw: bytes, config: dict) -> AcquireResult:
    """Parse Excel bytes into structured records with profiling.

    Security: uses read_only + data_only mode, enforces row/column limits,
    sanitizes cell values, and ensures workbook cleanup via try/finally.
    """
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException

    wb = None
    try:
        try:
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        except (InvalidFileException, Exception) as e:
            return AcquireResult(
                success=False, error=f"Invalid Excel file: {e}")

        sheet_names = wb.sheetnames
        sheet_name = config.get("sheet_name") or sheet_names[0]

        if sheet_name not in sheet_names:
            return AcquireResult(
                success=False, error=f"Sheet '{sheet_name}' not found. Available: {sheet_names}")

        ws = wb[sheet_name]

        # Read rows with limits to prevent memory exhaustion
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > MAX_ROWS:
                return AcquireResult(
                    success=False,
                    error=f"Sheet exceeds maximum row limit of {MAX_ROWS:,}.")
            rows.append(row)

        if not rows:
            return AcquireResult(success=True, record_count=0, metadata={"sheets": sheet_names})

        # Column count check
        if len(rows[0]) > MAX_COLUMNS:
            return AcquireResult(
                success=False,
                error=f"Sheet has {len(rows[0])} columns, exceeding limit of {MAX_COLUMNS}.")

        has_header = config.get("has_header", True)
        if has_header:
            headers = [
                _sanitize_header(str(h).strip()) if h is not None else f"column_{i}"
                for i, h in enumerate(rows[0])
            ]
            data_rows = rows[1:]
        else:
            headers = [f"column_{i}" for i in range(len(rows[0]))]
            data_rows = rows

        records = []
        for row_idx, row in enumerate(data_rows):
            record = {"_row_number": row_idx + 1}
            for col_idx, header in enumerate(headers):
                val = row[col_idx] if col_idx < len(row) else None
                # Convert datetime objects to ISO strings for JSON compatibility
                if isinstance(val, datetime):
                    val = val.isoformat()
                # Sanitize string values against formula injection
                elif isinstance(val, str):
                    val = _sanitize_cell(val)
                record[header] = val
            records.append(record)

        column_values = {h: [r.get(h) for r in records] for h in headers}
        schema_info = {
            "columns": [
                {"name": h, "type": _infer_column_type(column_values[h]), "index": i}
                for i, h in enumerate(headers)
            ],
            "sheet_name": sheet_name,
            "available_sheets": sheet_names,
        }

        profiling = {
            "row_count": len(records),
            "column_count": len(headers),
            "columns": {h: _profile_column(h, column_values[h]) for h in headers},
        }

        preview_rows = records[:MAX_PREVIEW_ROWS]

        return AcquireResult(
            success=True,
            record_count=len(records),
            records=records,
            schema_info=schema_info,
            profiling=profiling,
            preview_rows=preview_rows,
            metadata={"format": "xlsx", "sheet_name": sheet_name},
        )
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def _check_json_depth(obj: Any, max_depth: int = 20, current: int = 0) -> bool:
    """Check if a JSON object exceeds maximum nesting depth."""
    if current > max_depth:
        return False
    if isinstance(obj, dict):
        return all(_check_json_depth(v, max_depth, current + 1) for v in obj.values())
    if isinstance(obj, list):
        return all(_check_json_depth(v, max_depth, current + 1) for v in obj[:100])
    return True


def _sanitize_json_value(val: Any) -> Any:
    """Sanitize JSON values — prevent formula injection in string values."""
    if isinstance(val, str):
        return _sanitize_cell(val)
    if isinstance(val, dict):
        return {_sanitize_cell(str(k)) if isinstance(k, str) else k: _sanitize_json_value(v)
                for k, v in val.items()}
    if isinstance(val, list):
        return [_sanitize_json_value(v) for v in val]
    return val


def parse_json(raw: bytes, config: dict) -> AcquireResult:
    """Parse JSON/JSONL bytes into records.

    Security: enforces size limits, nesting depth limits, row limits,
    and sanitizes all string values against formula injection.
    """
    encoding = config.get("encoding") or _detect_encoding(raw)
    text = raw.decode(encoding, errors="replace").strip()

    # Size check for decoded text
    if len(text) > MAX_JSON_SIZE:
        return AcquireResult(
            success=False,
            error=f"JSON content exceeds {MAX_JSON_SIZE // (1024*1024)}MB limit.")

    records = []
    is_jsonl = config.get("jsonl", False)

    if is_jsonl or ("\n" in text and text.lstrip().startswith("{")):
        # JSONL: one JSON object per line
        lines = text.splitlines()
        if len(lines) > MAX_ROWS:
            return AcquireResult(
                success=False,
                error=f"JSONL file exceeds maximum row limit of {MAX_ROWS:,}.")
        for line_num, line in enumerate(lines, 1):
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
                if isinstance(obj, dict):
                    obj = _sanitize_json_value(obj)
                    obj["_row_number"] = line_num
                    records.append(obj)
            except json.JSONDecodeError:
                continue
    else:
        try:
            data = json.loads(text)
        except json.JSONDecodeError as e:
            return AcquireResult(success=False, error=f"Invalid JSON: {e}")

        # Depth check to prevent JSON bomb
        if not _check_json_depth(data):
            return AcquireResult(
                success=False,
                error="JSON exceeds maximum nesting depth of 20 levels.")

        if isinstance(data, list):
            if len(data) > MAX_ROWS:
                return AcquireResult(
                    success=False,
                    error=f"JSON array exceeds maximum row limit of {MAX_ROWS:,}.")
            for i, item in enumerate(data):
                if isinstance(item, dict):
                    item = _sanitize_json_value(item)
                    item["_row_number"] = i + 1
                    records.append(item)
                else:
                    records.append({"_row_number": i + 1, "value": item})
        elif isinstance(data, dict):
            # Check if any value is a list of dicts (common API pattern)
            array_key = config.get("records_path", "")
            if array_key and array_key in data:
                arr = data[array_key]
            else:
                # Auto-detect: find the first key whose value is a list of dicts
                arr = None
                for k, v in data.items():
                    if isinstance(v, list) and v and isinstance(v[0], dict):
                        arr = v
                        break
            if arr:
                if len(arr) > MAX_ROWS:
                    return AcquireResult(
                        success=False,
                        error=f"JSON array exceeds maximum row limit of {MAX_ROWS:,}.")
                for i, item in enumerate(arr):
                    if isinstance(item, dict):
                        item = _sanitize_json_value(item)
                        item["_row_number"] = i + 1
                        records.append(item)
            else:
                records.append({"_row_number": 1, **_sanitize_json_value(data)})

    if not records:
        return AcquireResult(success=True, record_count=0)

    # Collect all keys across all records for schema
    all_keys = []
    seen = set()
    for r in records:
        for k in r.keys():
            if k not in seen and k != "_row_number":
                all_keys.append(k)
                seen.add(k)

    column_values = {k: [r.get(k) for r in records] for k in all_keys}
    schema_info = {
        "columns": [
            {"name": k, "type": _infer_column_type(column_values[k]), "index": i}
            for i, k in enumerate(all_keys)
        ],
    }

    profiling = {
        "row_count": len(records),
        "column_count": len(all_keys),
        "columns": {k: _profile_column(k, column_values[k]) for k in all_keys},
    }

    return AcquireResult(
        success=True,
        record_count=len(records),
        records=records,
        schema_info=schema_info,
        profiling=profiling,
        preview_rows=records[:MAX_PREVIEW_ROWS],
        metadata={"format": "jsonl" if is_jsonl else "json"},
    )


def _tabulate(records: list[dict], metadata: dict) -> AcquireResult:
    """Schema, profiling and preview for an already-parsed record list.

    Shared by the readers that produce records directly rather than rows +
    header, so a new format gets the same profiling the CSV path has always
    produced instead of a thinner result that looks like a parsing failure.
    """
    all_keys: list[str] = []
    seen: set[str] = set()
    for r in records:
        for k in r:
            if k not in seen and k != "_row_number":
                all_keys.append(k)
                seen.add(k)

    column_values = {k: [r.get(k) for r in records] for k in all_keys}
    return AcquireResult(
        success=True,
        record_count=len(records),
        records=records,
        schema_info={
            "columns": [
                {"name": k, "type": _infer_column_type(column_values[k]), "index": i}
                for i, k in enumerate(all_keys)
            ],
        },
        profiling={
            "row_count": len(records),
            "column_count": len(all_keys),
            "columns": {k: _profile_column(k, column_values[k]) for k in all_keys},
        },
        preview_rows=records[:MAX_PREVIEW_ROWS],
        metadata=metadata,
    )

# ---------------------------------------------------------------------------
# XML
# ---------------------------------------------------------------------------

def _xml_repeating_element(root: "ET.Element") -> tuple[str, list]:
    """The element that repeats to form rows, and those rows.

    Government feeds nest the records one or two levels down —
    `<sdnList><sdnEntry>...` for OFAC, `<export><sanctionEntity>...` for the EU
    consolidated list — so the row element is rarely the document root's direct
    child by name. Pick the tag with the most siblings anywhere shallow, which
    is what "the repeating record" means without hardcoding a schema per feed.
    """
    best_tag, best_rows = "", []
    # Root's children, then grandchildren: deep enough for the real feeds,
    # shallow enough not to wander into per-record sub-structure.
    for container in [root, *list(root)]:
        counts: dict[str, list] = {}
        for child in container:
            counts.setdefault(_xml_localname(child.tag), []).append(child)
        for tag, rows in counts.items():
            if len(rows) > len(best_rows):
                best_tag, best_rows = tag, rows
    return best_tag, best_rows


def _xml_localname(tag: str) -> str:
    """Strip the namespace: `{urn:...}sanctionEntity` -> `sanctionEntity`."""
    return tag.rsplit("}", 1)[-1] if "}" in tag else tag


def _xml_flatten(elem: "ET.Element", prefix: str = "", depth: int = 0) -> dict:
    """One XML element to one flat record.

    Nested elements become dotted keys and attributes become `@name`, so a
    record keeps the structure the feed encoded rather than collapsing to text.
    Repeated siblings are joined rather than overwriting each other — an entity
    with four aliases should not silently keep only the last.
    """
    out: dict = {}
    if depth > MAX_XML_DEPTH:
        return out

    for key, value in elem.attrib.items():
        out[f"{prefix}@{_xml_localname(key)}"] = _sanitize_cell(str(value))

    text = (elem.text or "").strip()
    if text:
        out[prefix.rstrip(".") or "value"] = _sanitize_cell(text)

    for child in elem:
        name = f"{prefix}{_xml_localname(child.tag)}"
        nested = _xml_flatten(child, f"{name}.", depth + 1)
        for k, v in nested.items():
            if k in out:
                # Repeated element: keep both rather than overwrite.
                out[k] = f"{out[k]}; {v}" if str(v) not in str(out[k]) else out[k]
            else:
                out[k] = v
    return out


def parse_xml(raw: bytes, config: dict) -> AcquireResult:
    """Parse XML bytes into flat records.

    Uses defusedxml, not the stdlib parser. Verified against this Python
    (3.12): `xml.etree.ElementTree` refuses external entities but happily
    expands internal ones, so a billion-laughs document parses and inflates.
    defusedxml refuses both.

    `record_path` names the repeating element when auto-detection picks the
    wrong one.
    """
    try:
        from defusedxml.ElementTree import fromstring as _safe_fromstring
    except ImportError:  # pragma: no cover - dependency is in the lock
        return AcquireResult(success=False, error="XML support requires defusedxml.")

    encoding = config.get("encoding") or _detect_encoding(raw)
    text = raw.decode(encoding, errors="replace").strip()
    if len(text) > MAX_JSON_SIZE:
        return AcquireResult(
            success=False,
            error=f"XML content exceeds {MAX_JSON_SIZE // (1024*1024)}MB limit.")

    try:
        root = _safe_fromstring(text)
    except Exception as e:
        # Covers EntitiesForbidden/DTDForbidden as well as malformed input;
        # the message names which so an analyst can tell a hostile file from a
        # broken one.
        return AcquireResult(success=False, error=f"Invalid or unsafe XML ({type(e).__name__}): {e}")

    wanted = config.get("record_path", "")
    if wanted:
        rows = [el for el in root.iter() if _xml_localname(el.tag) == wanted]
        row_tag = wanted
    else:
        row_tag, rows = _xml_repeating_element(root)

    if not rows:
        # A single-record document is legitimate; treat the root as the record.
        rows, row_tag = [root], _xml_localname(root.tag)

    if len(rows) > MAX_ROWS:
        return AcquireResult(
            success=False,
            error=f"XML document exceeds maximum row limit of {MAX_ROWS:,}.")

    records = []
    for i, el in enumerate(rows, 1):
        rec = _xml_flatten(el)
        if not rec:
            continue
        rec["_row_number"] = i
        records.append(rec)

    if not records:
        return AcquireResult(success=True, record_count=0, metadata={"format": "xml"})

    return _tabulate(records, {"format": "xml", "record_element": row_tag})


# ---------------------------------------------------------------------------
# GeoJSON
# ---------------------------------------------------------------------------

def _geo_coords(geometry: dict) -> list[tuple[float, float]]:
    """Every (lng, lat) pair in a geometry, whatever its nesting depth.

    Point, LineString, Polygon, the Multi* variants and GeometryCollection all
    bottom out in [lng, lat] pairs; walking to the numbers avoids a branch per
    type and handles types added later.
    """
    out: list[tuple[float, float]] = []

    def walk(node, depth: int = 0):
        if depth > 12 or len(out) > 200_000:
            return
        if (isinstance(node, (list, tuple)) and len(node) >= 2
                and all(isinstance(c, (int, float)) for c in node[:2])):
            out.append((float(node[0]), float(node[1])))
            return
        if isinstance(node, (list, tuple)):
            for item in node:
                walk(item, depth + 1)

    if geometry.get("type") == "GeometryCollection":
        for g in geometry.get("geometries") or []:
            if isinstance(g, dict):
                out.extend(_geo_coords(g))
        return out
    walk(geometry.get("coordinates"))
    return out


def _geo_reduce(geometry: dict) -> dict:
    """Reduce a geometry to the point-and-bbox shape the platform stores.

    Locations are latitude/longitude scalars on a graph node — there is no
    geometry column and no PostGIS. Rather than refuse polygons, keep what the
    existing views can use: a representative point (so it plots, and
    /geo/within works), the bounding box (so containment can be answered
    without a spatial index), and the original geometry as text so nothing is
    lost if real geometry storage arrives later.
    """
    coords = _geo_coords(geometry)
    if not coords:
        return {}

    lngs = [c[0] for c in coords]
    lats = [c[1] for c in coords]
    # Mean of the vertices, not a true centroid: for an admin boundary or an
    # EEZ it lands inside the shape, and it is honest about being derived —
    # `geo_source` records that so it is never mistaken for a surveyed point.
    out = {
        "longitude": round(sum(lngs) / len(lngs), 6),
        "latitude": round(sum(lats) / len(lats), 6),
        "bbox_min_lng": round(min(lngs), 6),
        "bbox_min_lat": round(min(lats), 6),
        "bbox_max_lng": round(max(lngs), 6),
        "bbox_max_lat": round(max(lats), 6),
        "geometry_type": str(geometry.get("type", "")),
        "vertex_count": len(coords),
        "geo_source": "geojson",
    }
    if len(coords) == 1:
        # A Point is the point; nothing was reduced.
        out["geo_source"] = "geojson_point"
    return out


def parse_geojson(raw: bytes, config: dict) -> AcquireResult:
    """Parse GeoJSON into one record per feature.

    Feature properties become the record's columns, and the geometry is
    reduced to a point plus bounding box (see _geo_reduce) because that is
    what this platform can store today.
    """
    encoding = config.get("encoding") or _detect_encoding(raw)
    text = raw.decode(encoding, errors="replace").strip()
    if len(text) > MAX_JSON_SIZE:
        return AcquireResult(
            success=False,
            error=f"GeoJSON content exceeds {MAX_JSON_SIZE // (1024*1024)}MB limit.")

    try:
        data = json.loads(text)
    except json.JSONDecodeError as e:
        return AcquireResult(success=False, error=f"Invalid GeoJSON: {e}")
    if not isinstance(data, dict):
        return AcquireResult(success=False, error="GeoJSON must be an object.")

    gtype = data.get("type", "")
    if gtype == "FeatureCollection":
        features = data.get("features") or []
    elif gtype == "Feature":
        features = [data]
    elif gtype:
        # A bare geometry is valid GeoJSON, just propertyless.
        features = [{"type": "Feature", "properties": {}, "geometry": data}]
    else:
        return AcquireResult(success=False, error="Not GeoJSON: no 'type' member.")

    if len(features) > MAX_ROWS:
        return AcquireResult(
            success=False,
            error=f"GeoJSON exceeds maximum row limit of {MAX_ROWS:,}.")

    records = []
    skipped_no_geometry = 0
    for i, feat in enumerate(features, 1):
        if not isinstance(feat, dict):
            continue
        props = feat.get("properties")
        rec = dict(_sanitize_json_value(props)) if isinstance(props, dict) else {}

        geometry = feat.get("geometry")
        if isinstance(geometry, dict) and geometry.get("type"):
            rec.update(_geo_reduce(geometry))
        else:
            # Counted, not silently dropped: a feature collection that is all
            # attributes and no shapes is a different problem from a parse
            # failure, and the caller should be able to tell them apart.
            skipped_no_geometry += 1

        if feat.get("id") is not None and "id" not in rec:
            rec["id"] = _sanitize_json_value(feat["id"])
        rec["_row_number"] = i
        records.append(rec)

    if not records:
        return AcquireResult(success=True, record_count=0, metadata={"format": "geojson"})

    result = _tabulate(records, {
        "format": "geojson",
        "feature_count": len(records),
        "features_without_geometry": skipped_no_geometry,
    })
    return result



def detect_format(filename: str) -> str:
    """Detect file format from extension."""
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if ext in ("csv", "tsv"):
        return "csv"
    if ext in ("xlsx", "xls"):
        return "xlsx"
    if ext == "jsonl":
        return "jsonl"
    # Checked before plain json: a .geojson is JSON, but parsing it as JSON
    # would keep the geometry as an unusable nested blob instead of reducing it
    # to the point and bbox the map can plot.
    if ext == "geojson":
        return "geojson"
    if ext == "json":
        return "json"
    if ext in ("xml", "kml", "rss", "atom"):
        return "xml"
    return ext


@register_connector
class FlatFileConnector(SourceConnector):
    """Connector for structured flat files: CSV, Excel, JSON/JSONL."""

    source_type = "file_upload"
    config_keys = ()
    unattended = False
    capability_note = (
        "NOT collection the system performs. It is a request for the analyst to "
        "upload a file by hand, and stays at zero records until they do. Propose it "
        "only for material that genuinely cannot be fetched from the open web."
    )

    def configure(self, config: dict[str, Any]) -> dict[str, Any]:
        """Validate flat file config. Expects at minimum a file_format or filename."""
        file_format = config.get("file_format", "")
        filename = config.get("filename", "")

        if not file_format and filename:
            file_format = detect_format(filename)

        if file_format not in SUPPORTED_FORMATS and file_format not in ("", "auto"):
            raise ValueError(f"Unsupported format: {file_format}. Supported: {SUPPORTED_FORMATS}")

        return {
            "file_format": file_format,
            "filename": filename,
            "encoding": config.get("encoding", ""),
            "delimiter": config.get("delimiter", ""),
            "has_header": config.get("has_header", True),
            "sheet_name": config.get("sheet_name", ""),
            "records_path": config.get("records_path", ""),
        }

    async def test(self, config: dict[str, Any]) -> HealthStatus:
        """For file uploads, test is always healthy (files are provided at acquire time)."""
        return HealthStatus(status=ConnectorHealth.HEALTHY)

    async def discover(self, config: dict[str, Any]) -> dict[str, Any]:
        """Discover is handled during acquire for flat files."""
        return {"supported_formats": list(SUPPORTED_FORMATS)}

    async def acquire(self, config: dict[str, Any], since: datetime | None = None) -> AcquireResult:
        """Parse file bytes from config['file_bytes'].

        For flat file connector, the file bytes are provided directly
        (from an upload or a fetched URL). This connector does not
        fetch files itself — the collection plan orchestrator handles that.
        """
        file_bytes = config.get("file_bytes")
        if not file_bytes:
            return AcquireResult(success=False, error="No file_bytes provided")

        if isinstance(file_bytes, str):
            file_bytes = file_bytes.encode("utf-8")

        file_format = config.get("file_format", "")
        filename = config.get("filename", "")
        if not file_format:
            file_format = detect_format(filename)

        try:
            if file_format in ("csv", "tsv"):
                if file_format == "tsv":
                    config = {**config, "delimiter": "\t"}
                return parse_csv(file_bytes, config)
            elif file_format in ("xlsx", "xls"):
                return parse_excel(file_bytes, config)
            elif file_format == "jsonl":
                return parse_json(file_bytes, {**config, "jsonl": True})
            elif file_format == "geojson":
                return parse_geojson(file_bytes, config)
            elif file_format == "json":
                return parse_json(file_bytes, config)
            elif file_format == "xml":
                return parse_xml(file_bytes, config)
            else:
                return AcquireResult(
                    success=False, error=f"Unsupported format: {file_format}")
        except Exception as e:
            logger.exception("Flat file parse error: %s", e)
            return AcquireResult(success=False, error=str(e))
